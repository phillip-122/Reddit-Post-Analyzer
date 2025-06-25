'''
This Is my Reddit Post Analyzer. It analyzes post trends, performance, and engagement metrics across subreddits
The main tools that it utilizes is Pandas, Sentiment Analysis, Wordcloud, and MatPlotLib
'''


import requests
import pandas as pd 
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from wordcloud import WordCloud, STOPWORDS
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

def whatInfo():
    subreddit = input("Which subreddit would you like to do? r/").lower()

    print("What type of post do you want? ")
    print("1. New")
    print("2. Hot")
    print("3. Rising")
    print("4. Top")
    print("5. Controversial")
    postType = input("Enter a post type: ").lower()

    if postType == "top" or postType == "controversial":
        print("What time frame of data do you want? ")
        print("1. Hour")
        print("2. Day")
        print("3. Week")
        print("4. Month")
        print("5. Year")
        print("6. All")
        timeFrame = input("Enter a timeframe: ").lower()
    else:
        timeFrame = None
    
    compare = input("Would you like to compare this info to another subreddits? ").lower()

    if compare == "yes":
        secondSub = input("Which subreddit would you like to do? r/")
    else:
        secondSub = "" #The reason that we are doing an empty string instead of None is because when we go to print the output in fillExcel
                       #we need it to not show up as anything instead of printing the string "None"

    return subreddit, postType, timeFrame, secondSub

def setup(subreddit, postType, timeFrame=None):

    try:
        with open('client_ID.txt', 'r') as f:
            CLIENT_ID = f.read()
    except FileNotFoundError:
        print("Client ID file not found!")
        return
    
    try:
        with open('secret_KEY.txt', 'r') as f:
            SECRET_KEY = f.read()
    except FileNotFoundError:
        print("Secret Key file not found!")
        return

    try:
        with open('reddit_username.txt', 'r') as f:
            username = f.read()
    except FileNotFoundError:
        print("username file not found!")
        return
    
    try:
        with open('pw.txt', 'r') as f:
            pw = f.read()
    except FileNotFoundError:
        print("Password file not found!")
        return
    
    auth = requests.auth.HTTPBasicAuth(CLIENT_ID, SECRET_KEY)

    data = {'grant_type': 'password', 'username': username, 'password': pw}
    headers = {'User-Agent': 'myAPI/0.1'}

    res = requests.post('https://www.reddit.com/api/v1/access_token', auth=auth, data=data, headers=headers)
    
    if res.status_code != 200:
        print(f"Error: Unable to fetch access token. Status code {res.status_code}")
        return

    TOKEN = res.json()['access_token']

    headers['Authorization'] = f'bearer {TOKEN}'

    res = requests.get(f'https://oauth.reddit.com/r/{subreddit}/{postType}', headers=headers, params={'limit': 100, 't': timeFrame}) #if I want post after a specific date add  'after': 't3_1he8hbj' to params
                                                                                                      #we use the limit as 100 because reddit api does not allow you to grab more for free
    return res

def bestPostingTime(res):
    '''
    The purpose of this function is to take certain information out of the api, and then once it does that it converts the unix time to real world time
    in order for the use to be able to choose what timeframe they want to display.
    '''
    data = []
    for post in res.json()['data']['children']:
        data.append({
            'upvote_ratio': post['data']['upvote_ratio'],
            'score': post['data']['score'],
            'created_utc': post['data']['created_utc']
        })
    dfBestPostingTime = pd.DataFrame(data)

    dfBestPostingTime['created_time'] = pd.to_datetime(dfBestPostingTime['created_utc'], unit='s')
    dfBestPostingTime['hour_of_day'] = dfBestPostingTime['created_time'].dt.hour
    dfBestPostingTime['min_of_day'] = dfBestPostingTime['created_time'].dt.minute
    dfBestPostingTime['month'] = dfBestPostingTime['created_time'].dt.month
    dfBestPostingTime['year'] = dfBestPostingTime['created_time'].dt.year
    dfBestPostingTime['day_of_week'] = dfBestPostingTime['created_time'].dt.day_of_week
    dfBestPostingTime['day_name'] = dfBestPostingTime['created_time'].dt.day_name()

    dfBestPostingTime.drop('created_utc', axis=1, inplace=True)

    return dfBestPostingTime

def wordCloud(res, secondTime=None):
    '''
    The goal of this function is to take the title's of the posts and then it will extract the most commonly used words in order to determine which words
    may make a post more popular. Once it takes the title's and converts them into a wordcloud, it takes the wordcloud and finds how often the words
    appear in order to give you a list of the most popular words
    '''
    data = []

    for post in res.json()['data']['children']:
        data.append({
            'title': post['data']['title'],
        })
    dfWordCloud = pd.DataFrame(data)

    stopwords = STOPWORDS

    wc = WordCloud(background_color='white', stopwords=stopwords, height=600, width=400)
    text = ' '.join(dfWordCloud['title'])
    
    wc.generate(text)
    if secondTime == None:
        wc.to_file('wordcloud.png')
    else:
        wc.to_file('secondSub-WordCloud.png')

    totalNumWords = len(text.split())
    wordFreq = wc.words_
    absoluteFreq = {word: int(freq * totalNumWords) for word, freq in wordFreq.items()}

    dfWordCloud = pd.DataFrame(list(absoluteFreq.items()), columns=['Word', 'Absolute Frequency'])
    dfWordCloud['Rank'] = dfWordCloud['Absolute Frequency'].rank(ascending=False).astype(int)
    dfWordCloud.sort_values('Rank', inplace=True)

    return dfWordCloud

def sentimentAnalysis(res):
    '''
    The purpose of this function is to take the posts titles and selftext(selftext is just the way that the api defines the body text of a post)
    and then once it takes these 2 items, it will then use the VADER sentiment analyzer to determine if the post is more positive, negative,
    or neutral tone. This is useful because if a certain subreddit has a lot of posts that are either positive or negative, then that means that your post
    should be the same tone as well in order to maximize upvotes.
    '''
    data = []

    for post in res.json()['data']['children']:
        data.append({
            'title': post['data']['title'],
            'selftext': post['data']['selftext']
        })
    dfSentimentAnalysis = pd.DataFrame(data)

    analyzer = SentimentIntensityAnalyzer()

    dfSentimentAnalysis['sentiment title'] = dfSentimentAnalysis['title'].apply(lambda title: analyzer.polarity_scores(title)['compound'])
    dfSentimentAnalysis['sentiment selftext'] = dfSentimentAnalysis['selftext'].apply(lambda selftext: analyzer.polarity_scores(selftext)['compound'])
    #this combines the title and selftext and averages them out to see what the overall posts sentiment is
    dfSentimentAnalysis['Combined Sentiment'] = (dfSentimentAnalysis['sentiment title'] + dfSentimentAnalysis['sentiment selftext']) / 2
    dfSentimentAnalysis = dfSentimentAnalysis.sort_values(by='Combined Sentiment', ascending=False)

    #This takes the sentiment analysis then it takes the total number of posts, the number of positive, negative, and neutral posts
    #and then it converts it to a percent so that we can send it to the matplotlib and create a pie chart
    totalNum = 0
    numNegative = 0
    numPositive = 0
    numNeutral = 0
    
    pieChartData = []

    for index, row in dfSentimentAnalysis.iterrows():
        totalNum += 1
        
        if row['Combined Sentiment'] > 0:
            numPositive += 1
        elif row['Combined Sentiment'] < 0:
            numNegative += 1
        else:
            numNeutral += 1

    percentPos = (numPositive / totalNum) * 100
    percentNeg = (numNegative / totalNum) * 100
    percentNeu = (numNeutral / totalNum) * 100

    pieChartData.append(percentPos)
    pieChartData.append(percentNeg)
    pieChartData.append(percentNeu)

    return dfSentimentAnalysis, pieChartData

def plotInfo(dfBestPostingTime, pieChartData, dfBestPostingTime2=None, pieChartData2=None, subreddit=None, secondSub=""):
    #best posting time chart
    plt.figure(figsize=(12, 5))

    hourlyScores = dfBestPostingTime.groupby('hour_of_day')['score'].mean()

    if secondSub == "":
        plt.plot(hourlyScores.index, hourlyScores.values, marker='o', color='b')
        plt.legend([f"{subreddit}"])
    else:
        hourlyScores2 = dfBestPostingTime2.groupby('hour_of_day')['score'].mean()
        plt.plot(hourlyScores.index, hourlyScores.values, marker='o', color='b')
        plt.plot(hourlyScores2.index, hourlyScores2.values, marker='o', color='r')
        plt.legend([f"{subreddit}", f"{secondSub}"])

    plt.xlabel("Hour of the day")
    plt.ylabel("Score")
    plt.title("Best posting time", fontweight="bold")
    plt.xticks(range(24))
    plt.tight_layout()
    plt.savefig('best-upload-hour.png')
    plt.close()

    #sentiment analysis pie chart
    labels = ['postive', 'negative', 'neutral']

    if secondSub == "":
        plt.pie(pieChartData, autopct=lambda p: '{:.1f}%'.format(p) if p > 0 else '', textprops={'fontweight': 'bold'})
        plt.title(f"{subreddit}", fontweight="bold")
        plt.legend(labels)
        plt.tight_layout()
        plt.savefig('Sentiment-Analysis-Pie-Chart.png')
        plt.close()
    else:
        fig, axes = plt.subplots(1, 2, figsize=(10, 5))

        axes[0].pie(pieChartData, autopct=lambda p: '{:.1f}%'.format(p) if p > 0 else '', textprops={'fontweight': 'bold'})
        axes[0].set_title(f"{subreddit}", fontweight="bold")

        axes[1].pie(pieChartData2, autopct=lambda p: '{:.1f}%'.format(p) if p > 0 else '', textprops={'fontweight': 'bold'})
        axes[1].set_title(f"{secondSub}", fontweight="bold")

        fig.legend(labels)
        fig.tight_layout()
        fig.savefig('Sentiment-Analysis-Pie-Chart.png')
        plt.close(fig)

def removeImages(sheet):
    for image in sheet._images:
        sheet._images.remove(image)

def fillExcel(dfBestPostingTime, dfWordCloud, dfSentimentAnalysis, dfBestPostingTime2=None, dfWordCloud2=None, dfSentimentAnalysis2=None, subreddit="", secondSub=""):
    ptBestPostingTime = dfBestPostingTime.pivot_table(index=['hour_of_day'], values='score', aggfunc='mean')
    if secondSub == "":
        #Takes all the info and then puts it into an excel file, each on their own sheet
        with pd.ExcelWriter('pivot-table-reddit.xlsx', engine='openpyxl') as writer:
            ptBestPostingTime.to_excel(writer, sheet_name='Best Posting Time', startrow=1, startcol=0)
            dfWordCloud.to_excel(writer, sheet_name='WordCloud', startrow=1, startcol=0, index=False)
            dfSentimentAnalysis.to_excel(writer, sheet_name='Sentiment Analysis', startrow=1, startcol=0, index=False)

        #this takes all the images and puts them onto the excel file
        wb = load_workbook('pivot-table-reddit.xlsx')

        sheet = wb['WordCloud']
        removeImages(sheet)
        sheet['A1'] = (f"{subreddit}")
        sheet['A1'].font = Font(name='Times New Roman', bold=True, size=20)
        img = Image('wordcloud.png')
        sheet.add_image(img, 'G4')
    else:
        ptBestPostingTime2 = dfBestPostingTime2.pivot_table(index=['hour_of_day'], values='score', aggfunc='mean')

        with pd.ExcelWriter('reddit-project-output.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            ptBestPostingTime.to_excel(writer, sheet_name='Best Posting Time', startrow=1, startcol=0)
            ptBestPostingTime2.to_excel(writer, sheet_name='Best Posting Time', startrow=1, startcol=3)

            dfWordCloud.to_excel(writer, sheet_name='WordCloud', startrow=1, startcol=0, index=False)
            dfWordCloud2.to_excel(writer, sheet_name='WordCloud', startrow=1, startcol=4, index=False)

            dfSentimentAnalysis.to_excel(writer, sheet_name='Sentiment Analysis', startrow=1, startcol=0, index=False)
            dfSentimentAnalysis2.to_excel(writer, sheet_name='Sentiment Analysis', startrow=1, startcol=6, index=False)

        wb = load_workbook('reddit-project-output.xlsx')

        sheet = wb['WordCloud']
        removeImages(sheet)
        sheet['A1'] = (f"{subreddit}")
        sheet['A1'].font = Font(name='Times New Roman', bold=True, size=20)
        img = Image('wordcloud.png')
        sheet.add_image(img, 'I4')

        sheet = wb['WordCloud']
        sheet['E1'] = (f"{secondSub}")
        sheet['E1'].font = Font(name='Times New Roman', bold=True, size=20)
        img = Image('secondSub-WordCloud.png')
        sheet.add_image(img, 'P4')

    #We remove images because if you do not remove them then they will not get replaced by new ones, just overlapped
    sheet = wb['Best Posting Time']
    removeImages(sheet)
    sheet['A1'] = (f"{subreddit}")
    sheet['A1'].font = Font(name='Times New Roman', bold=True, size=20)
    sheet['D1'] = (f"{secondSub}")
    sheet['D1'].font = Font(name='Times New Roman', bold=True, size=20)

    img = Image('best-upload-hour.png')
    sheet.add_image(img, 'G4')

    sheet = wb['Sentiment Analysis']
    removeImages(sheet)
    sheet['A1'] = (f"{subreddit}")
    sheet['A1'].font = Font(name='Times New Roman', bold=True, size=20)
    sheet['G1'] = (f"{secondSub}")
    sheet['G1'].font = Font(name='Times New Roman', bold=True, size=20)
    img = Image('Sentiment-Analysis-Pie-Chart.png')
    sheet.add_image(img, 'P4')

    wb.save('reddit-project-output.xlsx')
    
def main():
    subreddit, postType, timeFrame, secondSub = whatInfo()
    res = setup(subreddit, postType, timeFrame)

    dfBestPostingTime = bestPostingTime(res)
    dfWordCloud = wordCloud(res)
    dfSentimentAnalysis, pieChartData = sentimentAnalysis(res)
    
    if secondSub != "":
        res2 = setup(secondSub, postType, timeFrame)
        dfBestPostingTime2 = bestPostingTime(res2)
        dfWordCloud2 = wordCloud(res2, secondTime='yes')
        dfSentimentAnalysis2, pieChartData2 = sentimentAnalysis(res2)
    else:
        dfBestPostingTime2 = None
        dfSentimentAnalysis2 = None
        dfWordCloud2 = None
        pieChartData2 = None

    plotInfo(dfBestPostingTime, pieChartData, dfBestPostingTime2, pieChartData2, subreddit, secondSub)

    fillExcel(dfBestPostingTime, dfWordCloud, dfSentimentAnalysis, dfBestPostingTime2, dfWordCloud2, dfSentimentAnalysis2, subreddit, secondSub)
    
if __name__ == "__main__":
    main()


